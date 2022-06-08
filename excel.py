import os
import pylightxl as xl
from popups import display_message
from openpyxl import load_workbook
from datetime import datetime, date
from win32com.client import DispatchEx

class ExcelFile():
    def __init__(self,id_numbers, office, directory):
        self.office = office
        self.directory = directory
        self.id_numbers = id_numbers

    def create_file(self):
        try:
            count = 10
            datestamp_ = date.today()
            timestamp_ = int(datetime.timestamp(datetime.now()))
            file_name = f"MHAI_{self.office}_{datestamp_}_{timestamp_}.xlsx"
            destination = os.path.join(self.directory, file_name).replace('\\','/')
            # wb = load_workbook(filename = 'template.xlsx')
            wb = xl.readxl(fn='template.xlsx')
            # ws = wb.active
            for id_number in self.id_numbers:
                # ws[f"B{count}"] = id_number
                # ws[f"C{count}"] = f"=LEFT({id_number},11)"
                wb.ws(ws='Duplicate ID').update_address(address=f"B{count}", val=id_number)
                count += 1
            # wb.save(filename = destination)
            # xl = DispatchEx('Excel.Application')
            # xl.Visible = False
            # updated_wb = xl.Workbooks.Open(destination)
            # updated_wb.Save()
            # updated_wb.Close()
            xl.writexl(db=wb, fn=destination)
            display_message(1, file_name=file_name, path=self.directory)
        except Exception as e:
            display_message(repr(e))